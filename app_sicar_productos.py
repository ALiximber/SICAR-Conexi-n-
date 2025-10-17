"""
App Sicar Productos 2sql — Excel ⇄ MySQL (estandarizada)

Cambios clave para portabilidad:
- Sin rutas ni credenciales hardcodeadas. Lee .env o variables de entorno.
- Carpeta de datos por defecto: ~/Documents/SicarData (editable desde la UI).
- Dependencias fijadas en requirements.txt.
- Modo "2 sentencias" garantizado en SUBIDA: 1) SELECT catálogo de unidades, 2) UPDATE masivo.
- Compatibilidad Windows / macOS / Linux.
"""

import sys
import os
import re
import math
import logging
from dataclasses import dataclass
from typing import Optional, Tuple, Dict, List

import numpy as np
import pandas as pd

# GUI
from PyQt5.QtWidgets import (
    QApplication, QWidget, QTabWidget, QVBoxLayout, QPushButton,
    QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QScrollArea,
    QSplitter, QLineEdit, QMessageBox, QSizePolicy, QStatusBar, QCheckBox, QSpacerItem, QFileDialog
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QPalette, QColor

# Carga .env si existe
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# === Atributos Qt ANTES de crear QApplication ===
QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

# ====== LOGGING ======
logging.basicConfig(
    level=os.getenv("APP_LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s"
)
log = logging.getLogger("sicar-app")

# ====== CONFIG MYSQL desde entorno ======
@dataclass
class MySQLConfig:
    host: str = os.getenv("MYSQL_HOST", "127.0.0.1")
    port: int = int(os.getenv("MYSQL_PORT", "3306"))
    user: str = os.getenv("MYSQL_USER", "root")
    password: str = os.getenv("MYSQL_PASSWORD", "")
    database: str = os.getenv("MYSQL_DB", "sicar")

MYSQL_CFG = MySQLConfig()

# ====== SQLAlchemy opcional ======
try:
    from sqlalchemy import create_engine
    from urllib.parse import quote_plus

    def make_sa_engine(cfg: MySQLConfig):
        url = (
            f"mysql+pymysql://{cfg.user}:{quote_plus(cfg.password)}"
            f"@{cfg.host}:{cfg.port}/{cfg.database}?charset=utf8mb4"
        )
        return create_engine(url, pool_recycle=3600, pool_pre_ping=True)
except Exception:
    make_sa_engine = None

# ====== CONSULTAS ======
SQL_PRODUCTOS = """
SELECT 
  a.art_id            AS id,
  a.clave             AS clave,
  a.clavealterna      AS clave_alterna,
  a.descripcion       AS descripcion,
  a.existencia        AS existencia,
  a.servicio          AS servicio,
  a.preciocompra      AS precio_compra,
  a.precio1           AS precio1,
  a.precio2           AS precio2,
  a.precio3           AS precio3,
  a.unidadCompra      AS unidad_compra,
  a.unidadVenta       AS unidad_venta,
  a.factor            AS factor,
  a.granel            AS granel
FROM articulo a
WHERE a.status=1
"""

SQL_LOOKUPS = """
SELECT 
  u.uni_id,
  UPPER(REPLACE(REPLACE(u.nombre,'.',''),' ','')) AS key_nombre,
  NULL AS key_abrev
FROM unidad u
"""

# ====== MAPEO EXCEL → BD Y TIPADO ======
ALLOWED_MAP = {
    'descripcion': ('descripcion', 'str'),
    'existencia': ('existencia', 'num'),
    'servicio': ('servicio', 'bool'),
    'precio_compra': ('preciocompra', 'num'),
    'precio1': ('precio1', 'num'),
    'precio2': ('precio2', 'num'),
    'precio3': ('precio3', 'num'),
    'unidad_compra': ('unidadCompra', 'unit'),
    'unidad_venta': ('unidadVenta', 'unit'),
    'factor': ('factor', 'num'),
    'granel': ('granel', 'bool'),
    'clave_alterna': ('clavealterna', 'str'),
}

UNIT_NORMALIZATION = {
    'PZ':'PIEZA','PZA':'PIEZA','PZAS':'PIEZA','PIEZA':'PIEZA','PIEZAS':'PIEZA',
    'KG':'KG','KILO':'KG','KILOGRAMO':'KG','KILOS':'KG','KGS':'KG',
    'L':'L','LT':'L','LTS':'L','LITRO':'L','LITROS':'L',
    'M':'M','MT':'M','MTS':'M','METRO':'M','METROS':'M',
    'CJ':'CAJA','CAJA':'CAJA','CAJ':'CAJA',
    'PAQ':'PAQUETE','PAQUETE':'PAQUETE',
}

# ===== Helpers de conversión =====
def _to_py_mysql(v):
    if v is None:
        return None
    if isinstance(v, (np.bool_, bool)):
        return int(bool(v))
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating, float)):
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    if isinstance(v, str):
        s = v.strip()
        return None if s == '' else s
    return v

def df_to_mysql_params(df: pd.DataFrame) -> list:
    obj = df.astype(object)
    obj = obj.where(pd.notnull(obj), None)
    recs = obj.to_dict(orient='records')
    return [{k: _to_py_mysql(v) for k, v in row.items()} for row in recs]

def norm_key(s: str) -> str:
    return re.sub(r"[ .]", "", str(s or '').upper())

# ===== UI =====
class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Excel ⇄ MySQL Productos — 2 SQL por subida')
        self.resize(1180, 760)

        default_data = os.path.join(os.path.expanduser('~'), 'Documents', 'SicarData')
        os.makedirs(default_data, exist_ok=True)
        self.data_folder = os.getenv("DATA_FOLDER", default_data)

        self.max_preview_rows = int(os.getenv("MAX_PREVIEW_ROWS", "50000"))
        self._tables: Dict[str, QTableWidget] = {}

        self.initUI()
        self.setStyleSheet(self.get_styles())

    def initUI(self):
        root = QVBoxLayout(self)

        top_bar = QHBoxLayout()
        title = QLabel('Vista previa de archivos Excel')
        title.setObjectName('titleLabel')
        title.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText('Filtrar filas en la pestaña actual…')
        self.search_box.textChanged.connect(self.apply_filter_current_tab)
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setMinimumWidth(600)
        self.search_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.load_all_chk = QCheckBox('Cargar todo')

        self.refresh_button = QPushButton('Actualizar pestañas')
        self.refresh_button.setObjectName('refreshButton')
        self.refresh_button.clicked.connect(self.update_tabs)

        self.folder_btn = QPushButton('Carpeta…')
        self.folder_btn.clicked.connect(self.choose_folder)

        top_bar.addWidget(title)
        top_bar.addStretch(1)
        top_bar.addWidget(self.folder_btn)
        top_bar.addWidget(self.search_box, 2)
        top_bar.addWidget(self.load_all_chk)
        top_bar.addWidget(self.refresh_button)

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(lambda _: self.apply_filter_current_tab())

        self.status = QStatusBar()
        self.status.setSizeGripEnabled(False)

        root.addLayout(top_bar)
        root.addWidget(self.tabs)
        root.addWidget(self.status)

        self.update_tabs()

    def choose_folder(self):
        path = QFileDialog.getExistingDirectory(self, 'Seleccionar carpeta', self.data_folder or os.path.expanduser('~'))
        if path:
            self.data_folder = path
            self.update_tabs()

    def update_tabs(self):
        self.tabs.clear()
        self._tables.clear()

        if not os.path.isdir(self.data_folder):
            self._empty_state(f"La carpeta no existe: {self.data_folder}")
            return

        files = [f for f in os.listdir(self.data_folder) if f.lower().endswith(('.xlsx', '.xls'))]
        if not files:
            self._empty_state('No se encontraron archivos .xlsx o .xls')
            return

        for file in sorted(files):
            self.tabs.addTab(self.create_tab(file), file)

        self.status.showMessage(f"{len(files)} archivo(s) — {self.data_folder}")

    def _empty_state(self, msg):
        w = QWidget()
        lay = QVBoxLayout(w)
        lab = QLabel(msg)
        lab.setAlignment(Qt.AlignCenter)
        lab.setObjectName('emptyLabel')
        lay.addStretch(1)
        lay.addWidget(lab)
        lay.addStretch(1)
        self.tabs.addTab(w, '—')
        self.status.showMessage(msg)

    def create_tab(self, filename):
        tab = QWidget()
        tab_layout = QVBoxLayout(tab)
        tab_layout.setContentsMargins(0, 0, 0, 0)
        tab_layout.setSpacing(0)

        header = QLabel(f'Vista previa de: {filename}')
        header.setObjectName('fileLabel')
        header.setAlignment(Qt.AlignCenter)
        header.setFixedHeight(28)
        tab_layout.addWidget(header)

        splitter = QSplitter(Qt.Horizontal)

        # --- Columna izquierda ---
        left_wrap = QWidget()
        left_col = QVBoxLayout(left_wrap)
        left_col.setContentsMargins(0, 0, 0, 0)
        left_col.setSpacing(0)

        preview_table = QTableWidget()
        preview_table.setAlternatingRowColors(True)
        preview_table.setSortingEnabled(True)
        preview_table.setSelectionBehavior(QTableWidget.SelectRows)
        preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        preview_table.setSizeAdjustPolicy(QTableWidget.AdjustToContents)
        preview_table.setWordWrap(False)
        preview_table.horizontalHeader().setStretchLastSection(True)
        preview_table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        table_font = QFont('Segoe UI', 12)
        header_font = QFont('Segoe UI', 13, QFont.DemiBold)
        preview_table.setFont(table_font)
        preview_table.horizontalHeader().setFont(header_font)
        preview_table.verticalHeader().setDefaultSectionSize(26)

        pal = preview_table.palette()
        pal.setColor(QPalette.Base, QColor('#0f1320'))
        pal.setColor(QPalette.AlternateBase, QColor('#141a2b'))
        pal.setColor(QPalette.Text, QColor('#e6e6e9'))
        preview_table.setPalette(pal)

        file_path = os.path.join(self.data_folder, filename)
        df, total_rows = self._read_excel_safe(file_path)
        limit = None if self.load_all_chk.isChecked() else self.max_preview_rows
        df_preview = df.head(limit) if limit and len(df) > limit else df
        truncated = len(df) > len(df_preview)

        self._fill_table(preview_table, df_preview)
        self._tables[filename] = preview_table

        scroll_area = QScrollArea()
        scroll_area.setWidget(preview_table)
        scroll_area.setWidgetResizable(True)
        left_col.addWidget(scroll_area)

        info = QLabel(self._truncate_msg(truncated, len(df_preview), total_rows))
        info.setObjectName('infoLabel')
        info.setFixedHeight(22)
        left_col.addWidget(info)

        splitter.addWidget(left_wrap)

        # --- Columna derecha ---
        right_wrap = QWidget()
        right_col = QVBoxLayout(right_wrap)
        right_col.setContentsMargins(8, 0, 8, 0)
        right_col.setSpacing(10)

        btn_descargar = QPushButton('Descargar base de datos')
        btn_descargar.setObjectName('exportButton')
        btn_descargar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        btn_descargar.clicked.connect(lambda: self.download_db_to_excel(filename))

        btn_subir = QPushButton('Subir archivo')
        btn_subir.setObjectName('importButton')
        btn_subir.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        btn_subir.clicked.connect(lambda: self.upload_excel_to_db(filename))

        right_col.addWidget(btn_descargar, 1)
        right_col.addWidget(btn_subir, 1)
        right_col.addItem(QSpacerItem(0, 0, QSizePolicy.Minimum, QSizePolicy.Expanding))

        splitter.addWidget(right_wrap)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 1)

        tab_layout.addWidget(splitter)
        return tab

    # ===== Helpers Excel/UI =====
    def _read_excel_safe(self, file_path: str) -> Tuple[pd.DataFrame, int]:
        try:
            df = pd.read_excel(file_path, engine="openpyxl")
            return df, len(df)
        except Exception as e:
            QMessageBox.critical(self, 'Error al leer', f'No se pudo leer:\n{file_path}\n\n{e}')
            log.exception("Lectura Excel falló")
            return pd.DataFrame(), 0

    def _fill_table(self, table: QTableWidget, df: pd.DataFrame):
        table.clear()
        if df.empty:
            table.setRowCount(0)
            table.setColumnCount(0)
            return
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        for col, name in enumerate(df.columns):
            table.setHorizontalHeaderItem(col, QTableWidgetItem(str(name)))
            for row in range(df.shape[0]):
                val = '' if pd.isna(df.iloc[row, col]) else str(df.iloc[row, col])
                item = QTableWidgetItem(val)
                item.setForeground(QColor('#e6e6e9'))
                table.setItem(row, col, item)
        table.resizeColumnsToContents()

    def _truncate_msg(self, truncated: bool, shown: int, total: int) -> str:
        if total == 0:
            return 'Sin datos en la hoja activa.'
        if truncated:
            return f'Mostrando {shown:n} de {total:n} filas. Activa "Cargar todo" para ver todas.'
        return f'Total filas: {total:n}.'

    def apply_filter_current_tab(self):
        idx = self.tabs.currentIndex()
        if idx < 0:
            return
        filename = self.tabs.tabText(idx)
        table = self._tables.get(filename)
        if not table:
            return
        text = self.search_box.text().strip().lower()
        rows = table.rowCount()
        cols = table.columnCount()
        if not text:
            for r in range(rows):
                table.setRowHidden(r, False)
            self.status.clearMessage()
            return
        matches = 0
        for r in range(rows):
            show = False
            for c in range(cols):
                item = table.item(r, c)
                if item and text in item.text().lower():
                    show = True
                    break
            table.setRowHidden(r, not show)
            if show:
                matches += 1
        self.status.showMessage(f'{matches:n} fila(s) coinciden con "{text}"')

    # ===== DB → Excel =====
    def download_db_to_excel(self, filename: str):
        file_path = os.path.join(self.data_folder, filename)
        if not file_path.lower().endswith('.xlsx'):
            QMessageBox.warning(self, 'Formato no soportado', 'Descarga solo escribe en .xlsx.')
            return
        try:
            if make_sa_engine is not None:
                engine = make_sa_engine(MYSQL_CFG)
                try:
                    df = pd.read_sql(SQL_PRODUCTOS, engine)
                finally:
                    engine.dispose()
            else:
                import pymysql
                conn = pymysql.connect(
                    host=MYSQL_CFG.host, port=MYSQL_CFG.port, user=MYSQL_CFG.user,
                    password=MYSQL_CFG.password, database=MYSQL_CFG.database,
                    cursorclass=pymysql.cursors.DictCursor
                )
                try:
                    df = pd.read_sql(SQL_PRODUCTOS, conn)
                finally:
                    conn.close()
            df = df.loc[:, ~df.columns.duplicated()]
            self._write_df_to_excel_file(file_path, df, font_size=12)
            table = self._tables.get(filename)
            if table is not None:
                self._fill_table(table, df)
            self.status.showMessage(f'Descargados {len(df):n} productos a {os.path.basename(file_path)}', 6000)
        except Exception as e:
            log.exception("Descarga falló")
            QMessageBox.critical(self, 'Error al descargar', f'Fallo al consultar/escribir.\n\n{e}')

    # ===== Excel → DB =====
    def upload_excel_to_db(self, filename: str):
        file_path = os.path.join(self.data_folder, filename)
        if not file_path.lower().endswith('.xlsx'):
            QMessageBox.warning(self, 'Formato no soportado', 'Subida espera un .xlsx con encabezados en la primera fila.')
            return
        df, _ = self._read_excel_safe(file_path)
        if df.empty:
            QMessageBox.warning(self, 'Excel vacío', 'No hay filas para subir.')
            return

        df.columns = [str(c).strip().lower() for c in df.columns]
        df = df.loc[:, ~df.columns.duplicated()]

        key_col = 'id' if 'id' in df.columns else ('clave' if 'clave' in df.columns else None)
        if key_col is None:
            QMessageBox.critical(self, 'Sin identificador', 'El Excel debe contener "id" o "clave".')
            return

        clean_df, report = self._clean_autocorrect(df, key_col)

        try:
            import pymysql
            conn = pymysql.connect(
                host=MYSQL_CFG.host, port=MYSQL_CFG.port, user=MYSQL_CFG.user,
                password=MYSQL_CFG.password, database=MYSQL_CFG.database,
                cursorclass=pymysql.cursors.DictCursor, autocommit=False
            )
            try:
                # ========== SQL #1: SELECT lookups ==========
                with conn.cursor() as cur:
                    cur.execute(SQL_LOOKUPS)
                    rows_lookup = cur.fetchall()

                unit_key_to_id: Dict[str, int] = {}
                for r in rows_lookup:
                    uid = r['uni_id']
                    if r['key_nombre']:
                        unit_key_to_id[r['key_nombre']] = uid
                    if r.get('key_abrev'):
                        unit_key_to_id[r['key_abrev']] = uid

                def to_unit_id(x):
                    if pd.isna(x) or str(x).strip() == '':
                        return None
                    s = str(x).strip().upper()
                    s = UNIT_NORMALIZATION.get(s, s)
                    key = re.sub(r"[ .]", "", s)
                    return unit_key_to_id.get(key)

                for col in ('unidad_compra', 'unidad_venta'):
                    if col in clean_df.columns:
                        clean_df[col] = clean_df[col].apply(to_unit_id)

                update_cols = [c for c in clean_df.columns if c in ALLOWED_MAP]
                present = update_cols + [key_col]
                where_col = 'art_id' if key_col == 'id' else 'clave'

                set_parts = []
                for c in update_cols:
                    dbcol, _ = ALLOWED_MAP[c]
                    set_parts.append(f"{dbcol}=COALESCE(%({c})s,{dbcol})")
                if not set_parts:
                    QMessageBox.warning(self, 'Sin columnas de datos', 'No hay columnas de producto editables.')
                    conn.rollback()
                    conn.close()
                    return
                sql_upd = f"UPDATE articulo SET {', '.join(set_parts)} WHERE {where_col}=%({key_col})s"

                # ========== SQL #2: UPDATE masivo ==========
                rows = df_to_mysql_params(clean_df[present])
                with conn.cursor() as cur:
                    cur.executemany(sql_upd, rows)
                conn.commit()

                self.status.showMessage(f"Actualizados {len(rows):n} registro(s) usando {key_col}", 6000)
                msg = (
                    f"Filas usadas: {len(rows):n}\n"
                    f"Filas descartadas: {report['dropped_rows']:n}\n"
                    f"Num. convertidos: {report['num_coerced']:n}\n"
                    f"Booleanos convertidos: {report['bool_coerced']:n}\n"
                    f"Unidades normalizadas: {report['unit_normalized']:n}\n"
                )
                QMessageBox.information(self, 'Subida completada', msg)
            except Exception:
                conn.rollback()
                log.exception("Subida falló, rollback")
                raise
            finally:
                conn.close()
        except Exception as e:
            QMessageBox.critical(self, 'Error al subir', f'Fallo la actualización.\n\n{e}')

    # ---------- Autocorrección de tipos ----------
    def _clean_autocorrect(self, df: pd.DataFrame, key_col: str):
        df = df.copy()
        keep = [c for c in df.columns if c in ('id','clave') or c in ALLOWED_MAP]
        df = df[keep]

        df[key_col] = df[key_col].astype(str).str.strip()
        df = df[df[key_col].notna() & (df[key_col] != '')]

        mask_template = pd.Series(False, index=df.index)
        sample_cols = [c for c in df.columns if c != key_col]
        for c in sample_cols:
            mask_template |= (df[c].astype(str).str.strip().str.lower() == c)
        dropped_rows = int(mask_template.sum())
        if dropped_rows:
            df = df[~mask_template]

        num_coerced = 0
        bool_coerced = 0
        unit_normalized = 0

        def parse_num(x):
            nonlocal num_coerced
            if pd.isna(x):
                return None
            if isinstance(x, (int, float)):
                return float(x)
            s = str(x).strip()
            if s == '':
                return None
            s = s.replace(',', '.')
            m = re.search(r"-?\d+(?:\.\d+)?", s)
            if m:
                num_coerced += 1 if s != m.group(0) else 0
                try:
                    return float(m.group(0))
                except Exception:
                    return None
            return None

        def parse_bool(x):
            nonlocal bool_coerced
            if pd.isna(x):
                return None
            if isinstance(x, (int, float)):
                bool_coerced += 1 if (x not in (0,1)) else 0
                return 1 if float(x) != 0.0 else 0
            s = str(x).strip().lower()
            if s in ('1','true','t','si','sí','y','yes'):
                if s not in ('1','true'):
                    bool_coerced += 1
                return 1
            if s in ('0','false','f','no','n'):
                if s not in ('0','false'):
                    bool_coerced += 1
                return 0
            try:
                v = float(s.replace(',', '.'))
                bool_coerced += 1
                return 1 if v != 0 else 0
            except Exception:
                return None

        def normalize_unit_name(x):
            nonlocal unit_normalized
            if pd.isna(x):
                return None
            s = str(x).strip()
            if s == '':
                return None
            if re.fullmatch(r"\d+(?:[\.,]\d+)?", s):
                unit_normalized += 1
                return None
            key = s.upper().replace('.', '').replace(' ', '')
            canon = UNIT_NORMALIZATION.get(key, s.upper())
            if canon != s:
                unit_normalized += 1
            return canon

        for col, (dbcol, typ) in ALLOWED_MAP.items():
            if col not in df.columns:
                continue
            if typ == 'num':
                df[col] = df[col].apply(parse_num)
            elif typ == 'bool':
                df[col] = df[col].apply(parse_bool)
            elif typ == 'unit':
                df[col] = df[col].apply(normalize_unit_name)
            else:
                df[col] = df[col].apply(lambda v: None if (pd.isna(v) or str(v).strip()=='') else str(v).strip())

        report = {
            'dropped_rows': dropped_rows,
            'num_coerced': num_coerced,
            'bool_coerced': bool_coerced,
            'unit_normalized': unit_normalized,
        }
        return df, report

    # ===== Escritura Excel =====
    def _write_df_to_excel_file(self, file_path: str, df: pd.DataFrame, font_size: int = 12):
        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except Exception as e:
            QMessageBox.critical(self, 'Dependencia faltante', f'Requiere openpyxl (pip install openpyxl).\n\n{e}')
            return
        try:
            try:
                wb = load_workbook(file_path)
            except Exception:
                wb = Workbook()
            target = wb.sheetnames[0] if wb.sheetnames else 'Hoja1'
            if target in wb.sheetnames:
                ws_old = wb[target]
                wb.remove(ws_old)
                ws = wb.create_sheet(title=target, index=0)
            else:
                ws = wb.create_sheet(title=target, index=0)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            header_font = Font(size=font_size, bold=True)
            cell_font = Font(size=font_size)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(vertical='center')
            for r in ws.iter_rows(min_row=2):
                for c in r:
                    c.font = cell_font
                    c.alignment = Alignment(vertical='center')
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column_cells:
                    v = cell.value
                    if v is None:
                        continue
                    ln = len(str(v))
                    if ln > max_len:
                        max_len = ln
                letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[letter].width = min(max_len + 2, 60)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = 'A2'
            wb.save(file_path)
        except Exception as e:
            log.exception("Escritura Excel falló")
            QMessageBox.critical(self, 'Error al escribir', f'No se pudo escribir en el Excel.\n\n{e}')

    def get_styles(self):
        return """
        QWidget { background-color: #11131a; color: #e6e6e9; font-family: 'Segoe UI', sans-serif; font-size: 13px; }
        QLabel#titleLabel { font-size: 18px; font-weight: 700; color: #ffffff; }
        QLabel#fileLabel { font-size: 14px; font-weight: 700; color: #cfd1ff; }
        QLabel#emptyLabel { font-size: 14px; color: #b0b0b8; }
        QLabel#infoLabel { color: #9aa0a6; font-size: 12px; }

        QLineEdit { background: #171a23; border: 1px solid #2a2f3a; border-radius: 8px; padding: 8px 10px; color: #e6e6e9; font-size: 13px; }
        QLineEdit:focus { border: 1px solid #5b8cff; }

        QPushButton { background-color: #1a1f2b; color: #ffffff; border: 1px solid #2c3444; border-radius: 10px; padding: 14px; font-weight: 600; font-size: 14px; }
        QPushButton:hover { background-color: #232a3a; }
        QPushButton:pressed { background-color: #141925; }

        QPushButton#importButton { background-color: #ff3b30; border: 1px solid #ff5a52; color: #fff; }
        QPushButton#exportButton { background-color: #2563eb; border: 1px solid #3b82f6; color: #fff; }

        QTabWidget::pane { border: 1px solid #2c3444; border-radius: 12px; background: #141824; }
        QTabBar::tab { background: #1b2030; color: #e6e6e9; padding: 8px 16px; border-top-left-radius: 10px; border-top-right-radius: 10px; margin-right: 2px; }
        QTabBar::tab:selected { background: #2a3350; font-weight: 700; }

        QTableWidget { background-color: #0f1320; color: #e6e6e9; gridline-color: #2c3444; border: 1px solid #1d2433; border-radius: 8px; }
        QHeaderView::section { background-color: #1c2233; color: #e6e6e9; padding: 8px; border: 1px solid #2a3244; font-weight: 700; font-size: 13px; }
        QTableWidget::item { background-color: #0f1320; color: #e6e6e9; padding: 6px; font-size: 13px; }
        QTableWidget::item:!selected:alternate { background-color: #141a2b; }
        QTableWidget::item:selected { background-color: #3b82f6; color: #ffffff; }

        QStatusBar { background: #131826; color: #cdd1d8; border-top: 1px solid #1f2636; }
        """


def main():
    app = QApplication(sys.argv)
    app.setFont(QFont('Segoe UI', 11))
    ex = App()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
